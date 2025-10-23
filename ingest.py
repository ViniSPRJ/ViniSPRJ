import csv
import logging
import os
import tempfile
from typing import Dict, Iterator, List

logger = logging.getLogger(__name__)


class DataIngestor:
    """Stream rows from a CSV or XLSM file in fixed-size chunks."""

    def __init__(self, path: str, chunksize: int = 500_000):
        self.path = path
        self.chunksize = chunksize
        self._temp_file = None

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self._cleanup()

    def _cleanup(self):
        """Clean up temporary files if they exist."""
        if self._temp_file and os.path.exists(self._temp_file):
            try:
                os.unlink(self._temp_file)
                logger.info(f"Cleaned up temporary file: {self._temp_file}")
            except Exception as e:
                logger.warning(f"Failed to clean up temp file {self._temp_file}: {e}")
            finally:
                self._temp_file = None

    def load(self) -> Iterator[List[Dict[str, str]]]:
        # Validate file exists
        if not os.path.exists(self.path):
            raise FileNotFoundError(f"Input file not found: {self.path}")

        if not os.path.isfile(self.path):
            raise ValueError(f"Path is not a file: {self.path}")

        # Case-insensitive file type detection
        lower_path = self.path.lower()

        if lower_path.endswith('.csv'):
            logger.info(f"Processing CSV file: {self.path}")
            return self._read_csv(self.path)

        if lower_path.endswith('.xlsm'):
            logger.info(f"Processing XLSM file: {self.path}")
            try:
                csv_path = self._xlsm_to_csv(self.path)
                return self._read_csv(csv_path)
            except ImportError:
                raise ImportError(
                    "openpyxl is required to process XLSM files. "
                    "Install it with: pip install openpyxl"
                )

        raise ValueError(f'Unsupported file format: {self.path}. Supported formats: .csv, .xlsm')

    def _read_csv(self, path: str) -> Iterator[List[Dict[str, str]]]:
        try:
            with open(path, newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                chunk: List[Dict[str, str]] = []
                row_count = 0

                for row in reader:
                    chunk.append(row)
                    row_count += 1

                    if len(chunk) >= self.chunksize:
                        logger.debug(f"Yielding chunk of {len(chunk)} rows (total processed: {row_count})")
                        yield chunk
                        chunk = []

                if chunk:
                    logger.debug(f"Yielding final chunk of {len(chunk)} rows")
                    yield chunk

                logger.info(f"Completed processing {row_count} rows from {path}")
        except UnicodeDecodeError:
            # Retry with different encoding
            logger.warning(f"UTF-8 decoding failed, retrying with latin-1 encoding")
            with open(path, newline='', encoding='latin-1') as f:
                reader = csv.DictReader(f)
                chunk: List[Dict[str, str]] = []

                for row in reader:
                    chunk.append(row)
                    if len(chunk) >= self.chunksize:
                        yield chunk
                        chunk = []

                if chunk:
                    yield chunk

    def _xlsm_to_csv(self, xlsm_path: str) -> str:
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required to process XLSM files. "
                "Install it with: pip install openpyxl"
            )

        logger.info(f"Converting XLSM to CSV: {xlsm_path}")
        wb = openpyxl.load_workbook(xlsm_path, read_only=True, data_only=True)
        ws = wb.active

        # Create a proper temporary file
        fd, temp_path = tempfile.mkstemp(suffix='.csv', prefix='deepagent_')
        self._temp_file = temp_path

        try:
            with os.fdopen(fd, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                # Write header row
                writer.writerow([cell.value for cell in next(ws.iter_rows())])
                # Write data rows
                for row in ws.iter_rows(min_row=2, values_only=True):
                    writer.writerow(row)

            logger.info(f"Converted XLSM to temporary CSV: {temp_path}")
            return temp_path
        except Exception as e:
            # Clean up on error
            self._cleanup()
            raise
